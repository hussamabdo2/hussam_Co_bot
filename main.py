# ============================================================
# MEDICAL REPRESENTATIVE MANAGEMENT BOT
# Python + Telegram + PostgreSQL
# نسخة مفعلة بالكامل + Excel Export
# ============================================================

import os
import logging
from datetime import datetime, date, timedelta
from io import BytesIO

from dotenv import load_dotenv

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    BigInteger,
    String,
    Text,
    Date,
    DateTime,
    Float,
    Boolean,
    ForeignKey,
    UniqueConstraint,
    func,
    or_,
)
from sqlalchemy.orm import (
    declarative_base,
    relationship,
    sessionmaker,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)

from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

# ============================================================
# الإعدادات
# ============================================================

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "")

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/medical_rep_bot"
)

COMPANY_NAME = os.getenv(
    "COMPANY_NAME",
    "الشركة الدوائية"
)

ADMIN_IDS_TEXT = os.getenv("ADMIN_IDS", "")

ADMIN_IDS = []

for value in ADMIN_IDS_TEXT.split(","):
    value = value.strip()

    if value.isdigit():
        ADMIN_IDS.append(int(value))

if not BOT_TOKEN:
    raise ValueError(
        "BOT_TOKEN غير موجود في متغيرات البيئة."
    )

if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL غير موجود في Railway."
    )

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace(
        "postgres://",
        "postgresql://",
        1
    )

# ============================================================
# Logging
# ============================================================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

logger = logging.getLogger(__name__)

# ============================================================
# Database
# ============================================================

Base = declarative_base()

engine = create_engine(
    DATABASE_URL,
    pool_pre_ping=True,
    future=True
)

SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine
)

# ============================================================
# Roles
# ============================================================

ROLE_ADMIN = "admin"
ROLE_MANAGER = "manager"
ROLE_REPRESENTATIVE = "representative"
ROLE_WAREHOUSE = "warehouse"
ROLE_PHARMACY_OWNER = "pharmacy_owner"
ROLE_PHARMACIST = "pharmacist"

# ============================================================
# Doctor categories
# ============================================================

DOCTOR_CATEGORY_A = "A"
DOCTOR_CATEGORY_B = "B"
DOCTOR_CATEGORY_C = "C"
DOCTOR_CATEGORY_D = "D"

# ============================================================
# USERS
# ============================================================

class User(Base):

    __tablename__ = "users"

    id = Column(Integer, primary_key=True)

    telegram_id = Column(
        BigInteger,
        unique=True,
        nullable=False,
        index=True
    )

    full_name = Column(String(255), nullable=True)

    username = Column(String(255), nullable=True)

    phone = Column(String(50), nullable=True)

    role = Column(
        String(50),
        default=ROLE_REPRESENTATIVE
    )

    is_active = Column(
        Boolean,
        default=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )

    updated_at = Column(
        DateTime,
        default=datetime.utcnow,
        onupdate=datetime.utcnow
    )


# ============================================================
# SPECIALTIES
# ============================================================

class Specialty(Base):

    __tablename__ = "specialties"

    id = Column(Integer, primary_key=True)

    name = Column(
        String(255),
        unique=True,
        nullable=False
    )

    description = Column(
        Text,
        nullable=True
    )


# ============================================================
# HOSPITALS
# ============================================================

class Hospital(Base):

    __tablename__ = "hospitals"

    id = Column(Integer, primary_key=True)

    name = Column(
        String(255),
        nullable=False,
        index=True
    )

    hospital_type = Column(
        String(100),
        nullable=True
    )

    governorate = Column(
        String(100),
        nullable=True
    )

    city = Column(
        String(100),
        nullable=True
    )

    district = Column(
        String(100),
        nullable=True
    )

    address = Column(
        Text,
        nullable=True
    )

    phone = Column(
        String(50),
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )


# ============================================================
# CLINICS
# ============================================================

class Clinic(Base):

    __tablename__ = "clinics"

    id = Column(Integer, primary_key=True)

    name = Column(
        String(255),
        nullable=False
    )

    governorate = Column(
        String(100),
        nullable=True
    )

    city = Column(
        String(100),
        nullable=True
    )

    district = Column(
        String(100),
        nullable=True
    )

    address = Column(
        Text,
        nullable=True
    )

    phone = Column(
        String(50),
        nullable=True
    )


# ============================================================
# DOCTORS
# ============================================================

class Doctor(Base):

    __tablename__ = "doctors"

    id = Column(Integer, primary_key=True)

    full_name = Column(
        String(255),
        nullable=False,
        index=True
    )

    phone = Column(
        String(50),
        nullable=True,
        index=True
    )

    specialty_id = Column(
        Integer,
        ForeignKey("specialties.id"),
        nullable=True
    )

    category = Column(
        String(10),
        default=DOCTOR_CATEGORY_D
    )

    scientific_degree = Column(
        String(255),
        nullable=True
    )

    governorate = Column(
        String(100),
        nullable=True
    )

    city = Column(
        String(100),
        nullable=True
    )

    district = Column(
        String(100),
        nullable=True
    )

    address = Column(
        Text,
        nullable=True
    )

    working_days = Column(
        String(255),
        nullable=True
    )

    working_hours = Column(
        String(255),
        nullable=True
    )

    importance_score = Column(
        Float,
        default=0
    )

    prescription_score = Column(
        Float,
        default=0
    )

    last_visit_date = Column(
        Date,
        nullable=True
    )

    next_visit_date = Column(
        Date,
        nullable=True
    )

    assigned_rep_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    is_active = Column(
        Boolean,
        default=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )

    specialty = relationship("Specialty")

    assigned_rep = relationship(
        "User",
        foreign_keys=[assigned_rep_id]
    )


# ============================================================
# DOCTOR HOSPITAL
# ============================================================

class DoctorHospital(Base):

    __tablename__ = "doctor_hospitals"

    id = Column(Integer, primary_key=True)

    doctor_id = Column(
        Integer,
        ForeignKey("doctors.id"),
        nullable=False
    )

    hospital_id = Column(
        Integer,
        ForeignKey("hospitals.id"),
        nullable=False
    )

    department = Column(
        String(255),
        nullable=True
    )

    doctor = relationship("Doctor")

    hospital = relationship("Hospital")

    __table_args__ = (
        UniqueConstraint(
            "doctor_id",
            "hospital_id",
            name="uq_doctor_hospital"
        ),
    )


# ============================================================
# DOCTOR CLINIC
# ============================================================

class DoctorClinic(Base):

    __tablename__ = "doctor_clinics"

    id = Column(Integer, primary_key=True)

    doctor_id = Column(
        Integer,
        ForeignKey("doctors.id"),
        nullable=False
    )

    clinic_id = Column(
        Integer,
        ForeignKey("clinics.id"),
        nullable=False
    )

    doctor = relationship("Doctor")

    clinic = relationship("Clinic")


# ============================================================
# PHARMACIES
# ============================================================

class Pharmacy(Base):

    __tablename__ = "pharmacies"

    id = Column(Integer, primary_key=True)

    name = Column(
        String(255),
        nullable=False,
        index=True
    )

    owner_name = Column(
        String(255),
        nullable=True
    )

    phone = Column(
        String(50),
        nullable=True
    )

    governorate = Column(
        String(100),
        nullable=True
    )

    city = Column(
        String(100),
        nullable=True
    )

    district = Column(
        String(100),
        nullable=True
    )

    address = Column(
        Text,
        nullable=True
    )

    latitude = Column(
        Float,
        nullable=True
    )

    longitude = Column(
        Float,
        nullable=True
    )

    classification = Column(
        String(50),
        default="عادية"
    )

    assigned_rep_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    last_visit_date = Column(
        Date,
        nullable=True
    )

    next_visit_date = Column(
        Date,
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    is_active = Column(
        Boolean,
        default=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )

    assigned_rep = relationship(
        "User",
        foreign_keys=[assigned_rep_id]
    )


# ============================================================
# USER PHARMACY
# ============================================================

class UserPharmacy(Base):

    __tablename__ = "user_pharmacies"

    id = Column(Integer, primary_key=True)

    user_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=False
    )

    pharmacy_id = Column(
        Integer,
        ForeignKey("pharmacies.id"),
        nullable=False
    )

    user = relationship("User")

    pharmacy = relationship("Pharmacy")

    __table_args__ = (
        UniqueConstraint(
            "user_id",
            "pharmacy_id",
            name="uq_user_pharmacy"
        ),
    )


# ============================================================
# PHARMACISTS
# ============================================================

class Pharmacist(Base):

    __tablename__ = "pharmacists"

    id = Column(Integer, primary_key=True)

    full_name = Column(
        String(255),
        nullable=False
    )

    phone = Column(
        String(50),
        nullable=True
    )

    pharmacy_id = Column(
        Integer,
        ForeignKey("pharmacies.id"),
        nullable=True
    )

    classification = Column(
        String(100),
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    pharmacy = relationship("Pharmacy")


# ============================================================
# PRODUCTS
# ============================================================

class Product(Base):

    __tablename__ = "products"

    id = Column(Integer, primary_key=True)

    code = Column(
        String(100),
        unique=True,
        nullable=False,
        index=True
    )

    brand_name = Column(
        String(255),
        nullable=False,
        index=True
    )

    scientific_name = Column(
        String(255),
        nullable=True,
        index=True
    )

    active_ingredient = Column(
        String(255),
        nullable=True
    )

    concentration = Column(
        String(100),
        nullable=True
    )

    dosage_form = Column(
        String(100),
        nullable=True
    )

    package_size = Column(
        String(100),
        nullable=True
    )

    therapeutic_class = Column(
        String(255),
        nullable=True
    )

    manufacturer = Column(
        String(255),
        nullable=True
    )

    country_of_origin = Column(
        String(100),
        nullable=True
    )

    price = Column(
        Float,
        default=0
    )

    minimum_stock = Column(
        Integer,
        default=0
    )

    scientific_notes = Column(
        Text,
        nullable=True
    )

    image_file_id = Column(
        String(255),
        nullable=True
    )

    leaflet_file_id = Column(
        String(255),
        nullable=True
    )

    is_active = Column(
        Boolean,
        default=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )


# ============================================================
# PRODUCT SPECIALTY
# ============================================================

class ProductSpecialty(Base):

    __tablename__ = "product_specialties"

    id = Column(Integer, primary_key=True)

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    specialty_id = Column(
        Integer,
        ForeignKey("specialties.id"),
        nullable=False
    )

    product = relationship("Product")

    specialty = relationship("Specialty")


# ============================================================
# BATCHES
# ============================================================

class Batch(Base):

    __tablename__ = "batches"

    id = Column(Integer, primary_key=True)

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    batch_number = Column(
        String(100),
        nullable=False,
        index=True
    )

    manufacture_date = Column(
        Date,
        nullable=True
    )

    expiry_date = Column(
        Date,
        nullable=True,
        index=True
    )

    product = relationship("Product")

    __table_args__ = (
        UniqueConstraint(
            "product_id",
            "batch_number",
            name="uq_product_batch"
        ),
    )


# ============================================================
# WAREHOUSES
# ============================================================

class Warehouse(Base):

    __tablename__ = "warehouses"

    id = Column(Integer, primary_key=True)

    name = Column(
        String(255),
        nullable=False
    )

    governorate = Column(
        String(100),
        nullable=True
    )

    city = Column(
        String(100),
        nullable=True
    )

    address = Column(
        Text,
        nullable=True
    )

    manager_name = Column(
        String(255),
        nullable=True
    )

    phone = Column(
        String(50),
        nullable=True
    )

    is_active = Column(
        Boolean,
        default=True
    )


# ============================================================
# WAREHOUSE STOCK
# ============================================================

class WarehouseStock(Base):

    __tablename__ = "warehouse_stock"

    id = Column(Integer, primary_key=True)

    warehouse_id = Column(
        Integer,
        ForeignKey("warehouses.id"),
        nullable=False
    )

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    batch_id = Column(
        Integer,
        ForeignKey("batches.id"),
        nullable=True
    )

    quantity = Column(
        Integer,
        default=0
    )

    reserved_quantity = Column(
        Integer,
        default=0
    )

    updated_at = Column(
        DateTime,
        default=datetime.utcnow,
        onupdate=datetime.utcnow
    )

    warehouse = relationship("Warehouse")

    product = relationship("Product")

    batch = relationship("Batch")

    __table_args__ = (
        UniqueConstraint(
            "warehouse_id",
            "product_id",
            "batch_id",
            name="uq_warehouse_product_batch"
        ),
    )


# ============================================================
# PHARMACY STOCK
# ============================================================

class PharmacyStock(Base):

    __tablename__ = "pharmacy_stock"

    id = Column(Integer, primary_key=True)

    pharmacy_id = Column(
        Integer,
        ForeignKey("pharmacies.id"),
        nullable=False,
        index=True
    )

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    batch_id = Column(
        Integer,
        ForeignKey("batches.id"),
        nullable=True
    )

    quantity = Column(
        Integer,
        default=0
    )

    last_reported_by = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    updated_at = Column(
        DateTime,
        default=datetime.utcnow,
        onupdate=datetime.utcnow
    )

    pharmacy = relationship("Pharmacy")

    product = relationship("Product")

    batch = relationship("Batch")

    __table_args__ = (
        UniqueConstraint(
            "pharmacy_id",
            "product_id",
            "batch_id",
            name="uq_pharmacy_product_batch"
        ),
    )


# ============================================================
# STOCK MOVEMENTS
# ============================================================

class StockMovement(Base):

    __tablename__ = "stock_movements"

    id = Column(Integer, primary_key=True)

    location_type = Column(
        String(50),
        nullable=False
    )

    location_id = Column(
        Integer,
        nullable=False
    )

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    batch_id = Column(
        Integer,
        ForeignKey("batches.id"),
        nullable=True
    )

    movement_type = Column(
        String(100),
        nullable=False
    )

    quantity_before = Column(
        Integer,
        default=0
    )

    quantity_change = Column(
        Integer,
        default=0
    )

    quantity_after = Column(
        Integer,
        default=0
    )

    user_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )


# ============================================================
# DOCTOR VISITS
# ============================================================

class DoctorVisit(Base):

    __tablename__ = "doctor_visits"

    id = Column(Integer, primary_key=True)

    doctor_id = Column(
        Integer,
        ForeignKey("doctors.id"),
        nullable=False
    )

    representative_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    visit_date = Column(
        Date,
        default=date.today
    )

    interest_level = Column(
        String(100),
        nullable=True
    )

    notes = Column(
        Text,
        nullable=True
    )

    next_visit_date = Column(
        Date,
        nullable=True
    )

    doctor = relationship("Doctor")

    representative = relationship("User")


# ============================================================
# DOCTOR VISIT PRODUCTS
# ============================================================

class DoctorVisitProduct(Base):

    __tablename__ = "doctor_visit_products"

    id = Column(Integer, primary_key=True)

    visit_id = Column(
        Integer,
        ForeignKey("doctor_visits.id"),
        nullable=False
    )

    product_id = Column(
        Integer,
        ForeignKey("products.id"),
        nullable=False
    )

    interest_score = Column(
        Float,
        default=0
    )


# ============================================================
# PHARMACY VISITS
# ============================================================

class PharmacyVisit(Base):

    __tablename__ = "pharmacy_visits"

    id = Column(Integer, primary_key=True)

    pharmacy_id = Column(
        Integer,
        ForeignKey("pharmacies.id"),
        nullable=False
    )

    representative_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    visit_date = Column(
        Date,
        default=date.today
    )

    notes = Column(
        Text,
        nullable=True
    )

    next_visit_date = Column(
        Date,
        nullable=True
    )


# ============================================================
# VISIT REQUESTS
# ============================================================

class VisitRequest(Base):

    __tablename__ = "visit_requests"

    id = Column(Integer, primary_key=True)

    pharmacy_id = Column(
        Integer,
        ForeignKey("pharmacies.id"),
        nullable=False
    )

    requested_by_user_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=False
    )

    status = Column(
        String(50),
        default="pending"
    )

    notes = Column(
        Text,
        nullable=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )


# ============================================================
# REMINDERS
# ============================================================

class Reminder(Base):

    __tablename__ = "reminders"

    id = Column(Integer, primary_key=True)

    user_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=False
    )

    title = Column(
        String(255),
        nullable=False
    )

    message = Column(
        Text,
        nullable=True
    )

    reminder_date = Column(
        DateTime,
        nullable=False
    )

    is_sent = Column(
        Boolean,
        default=False
    )


# ============================================================
# AUDIT LOG
# ============================================================

class AuditLog(Base):

    __tablename__ = "audit_logs"

    id = Column(Integer, primary_key=True)

    user_id = Column(
        Integer,
        ForeignKey("users.id"),
        nullable=True
    )

    action = Column(
        String(255),
        nullable=False
    )

    entity_type = Column(
        String(100),
        nullable=True
    )

    entity_id = Column(
        Integer,
        nullable=True
    )

    old_value = Column(
        Text,
        nullable=True
    )

    new_value = Column(
        Text,
        nullable=True
    )

    created_at = Column(
        DateTime,
        default=datetime.utcnow
    )


# ============================================================
# DATABASE INIT
# ============================================================

def initialize_database():

    Base.metadata.create_all(
        bind=engine
    )


# ============================================================
# USER
# ============================================================

def get_or_create_user(update: Update):

    telegram_user = update.effective_user

    if not telegram_user:
        return None

    db = SessionLocal()

    try:

        user = db.query(User).filter(
            User.telegram_id == telegram_user.id
        ).first()

        if not user:

            role = ROLE_REPRESENTATIVE

            if telegram_user.id in ADMIN_IDS:
                role = ROLE_ADMIN

            user = User(
                telegram_id=telegram_user.id,
                full_name=telegram_user.full_name,
                username=telegram_user.username,
                role=role
            )

            db.add(user)

            db.commit()

            db.refresh(user)

        else:

            user.full_name = telegram_user.full_name
            user.username = telegram_user.username

            if (
                telegram_user.id in ADMIN_IDS
                and user.role != ROLE_ADMIN
            ):
                user.role = ROLE_ADMIN

            db.commit()

        return {
            "id": user.id,
            "telegram_id": user.telegram_id,
            "full_name": user.full_name,
            "role": user.role
        }

    finally:

        db.close()


def is_admin(user_data):

    return bool(
        user_data
        and user_data["role"] == ROLE_ADMIN
    )


# ============================================================
# AUDIT
# ============================================================

def add_audit_log(
    db,
    user_id,
    action,
    entity_type=None,
    entity_id=None,
    old_value=None,
    new_value=None
):

    db.add(
        AuditLog(
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            old_value=old_value,
            new_value=new_value
        )
    )


# ============================================================
# ADMIN KEYBOARD
# ============================================================

def admin_keyboard():

    keyboard = [

        [
            InlineKeyboardButton(
                "👨‍⚕️ الأطباء",
                callback_data="admin_doctors"
            ),
            InlineKeyboardButton(
                "➕ إضافة طبيب",
                callback_data="add_doctor"
            )
        ],

        [
            InlineKeyboardButton(
                "🏪 الصيدليات",
                callback_data="admin_pharmacies"
            ),
            InlineKeyboardButton(
                "➕ إضافة صيدلية",
                callback_data="add_pharmacy"
            )
        ],

        [
            InlineKeyboardButton(
                "💊 المنتجات",
                callback_data="admin_products"
            ),
            InlineKeyboardButton(
                "➕ إضافة منتج",
                callback_data="add_product"
            )
        ],

        [
            InlineKeyboardButton(
                "🧑‍🔬 الصيادلة",
                callback_data="admin_pharmacists"
            ),
            InlineKeyboardButton(
                "🏥 المستشفيات",
                callback_data="admin_hospitals"
            )
        ],

        [
            InlineKeyboardButton(
                "📦 المخازن",
                callback_data="admin_warehouses"
            ),
            InlineKeyboardButton(
                "👨‍💼 المستخدمون",
                callback_data="admin_users"
            )
        ],

        [
            InlineKeyboardButton(
                "📊 التقارير",
                callback_data="admin_reports"
            ),
            InlineKeyboardButton(
                "📈 الإحصائيات",
                callback_data="dashboard"
            )
        ],

        [
            InlineKeyboardButton(
                "⚠️ التنبيهات",
                callback_data="admin_alerts"
            ),
            InlineKeyboardButton(
                "🔎 البحث",
                callback_data="search"
            )
        ],

        [
            InlineKeyboardButton(
                "📥 تصدير Excel",
                callback_data="export_excel"
            )
        ],

    ]

    return InlineKeyboardMarkup(keyboard)


# ============================================================
# REPRESENTATIVE KEYBOARD
# ============================================================

def representative_keyboard():

    keyboard = [

        [
            InlineKeyboardButton(
                "👨‍⚕️ الأطباء",
                callback_data="rep_doctors"
            ),
            InlineKeyboardButton(
                "🏪 الصيدليات",
                callback_data="rep_pharmacies"
            )
        ],

        [
            InlineKeyboardButton(
                "🗓️ زيارات اليوم",
                callback_data="today_visits"
            ),
            InlineKeyboardButton(
                "➕ تسجيل زيارة",
                callback_data="add_visit"
            )
        ],

        [
            InlineKeyboardButton(
                "💊 المنتجات",
                callback_data="products_list"
            ),
            InlineKeyboardButton(
                "🔎 البحث",
                callback_data="search"
            )
        ],

        [
            InlineKeyboardButton(
                "🔔 التذكيرات",
                callback_data="my_reminders"
            )
        ],

        [
            InlineKeyboardButton(
                "📥 تصدير بياناتي Excel",
                callback_data="export_excel"
            )
        ]

    ]

    return InlineKeyboardMarkup(keyboard)


# ============================================================
# PHARMACY OWNER
# ============================================================

def pharmacy_owner_keyboard():

    keyboard = [

        [
            InlineKeyboardButton(
                "🏪 صيدليتي",
                callback_data="my_pharmacy"
            )
        ],

        [
            InlineKeyboardButton(
                "📦 مخزوني",
                callback_data="pharmacy_inventory"
            ),
            InlineKeyboardButton(
                "➕ إضافة صنف",
                callback_data="pharmacy_add_stock"
            )
        ],

        [
            InlineKeyboardButton(
                "✏️ تحديث كمية",
                callback_data="pharmacy_update_stock"
            ),
            InlineKeyboardButton(
                "📋 قائمة الأصناف",
                callback_data="pharmacy_products"
            )
        ],

        [
            InlineKeyboardButton(
                "⚠️ مخزون منخفض",
                callback_data="pharmacy_low_stock"
            ),
            InlineKeyboardButton(
                "⏳ قرب الانتهاء",
                callback_data="pharmacy_expiry"
            )
        ],

        [
            InlineKeyboardButton(
                "🗓️ طلب زيارة مندوب",
                callback_data="request_visit"
            ),
            InlineKeyboardButton(
                "💬 إرسال ملاحظة",
                callback_data="pharmacy_note"
            )
        ],

        [
            InlineKeyboardButton(
                "📥 تصدير مخزوني Excel",
                callback_data="export_excel"
            )
        ]

    ]

    return InlineKeyboardMarkup(keyboard)


# ============================================================
# MAIN KEYBOARD
# ============================================================

def get_main_keyboard(user):

    if user["role"] == ROLE_ADMIN:
        return admin_keyboard()

    if user["role"] in [
        ROLE_PHARMACY_OWNER,
        ROLE_PHARMACIST
    ]:
        return pharmacy_owner_keyboard()

    return representative_keyboard()


# ============================================================
# STATES
# ============================================================

(
    ADD_DOCTOR_NAME,
    ADD_DOCTOR_PHONE,
    ADD_DOCTOR_SPECIALTY,
    ADD_DOCTOR_CATEGORY,
    ADD_DOCTOR_CITY,

    ADD_PHARMACY_NAME,
    ADD_PHARMACY_OWNER,
    ADD_PHARMACY_PHONE,
    ADD_PHARMACY_CITY,

    ADD_PRODUCT_CODE,
    ADD_PRODUCT_NAME,
    ADD_PRODUCT_SCIENTIFIC,
    ADD_PRODUCT_CONCENTRATION,
    ADD_PRODUCT_FORM,

    ADD_STOCK_PRODUCT,
    ADD_STOCK_BATCH,
    ADD_STOCK_QUANTITY,
    ADD_STOCK_EXPIRY,

    SEARCH_TEXT,

) = range(19)


# ============================================================
# START
# ============================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user = get_or_create_user(update)

    text = (
        f"🩺 أهلاً بك في نظام {COMPANY_NAME}\n\n"
        "نظام إدارة الشركة الدوائية\n"
        "والأطباء والصيدليات والصيادلة والمنتجات "
        "والمخزون والزيارات والتقارير.\n\n"
        "اختر الخدمة من القائمة:"
    )

    await update.message.reply_text(
        text,
        reply_markup=get_main_keyboard(user)
    )


# ============================================================
# DASHBOARD
# ============================================================

async def dashboard_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        doctors = db.query(
            func.count(Doctor.id)
        ).scalar() or 0

        pharmacies = db.query(
            func.count(Pharmacy.id)
        ).scalar() or 0

        pharmacists = db.query(
            func.count(Pharmacist.id)
        ).scalar() or 0

        products = db.query(
            func.count(Product.id)
        ).scalar() or 0

        batches = db.query(
            func.count(Batch.id)
        ).scalar() or 0

        hospitals = db.query(
            func.count(Hospital.id)
        ).scalar() or 0

        warehouses = db.query(
            func.count(Warehouse.id)
        ).scalar() or 0

        users = db.query(
            func.count(User.id)
        ).scalar() or 0

        today = date.today()

        doctor_visits = db.query(
            func.count(DoctorVisit.id)
        ).filter(
            DoctorVisit.visit_date == today
        ).scalar() or 0

        pharmacy_visits = db.query(
            func.count(PharmacyVisit.id)
        ).filter(
            PharmacyVisit.visit_date == today
        ).scalar() or 0

        text = (
            "📈 لوحة الإحصائيات\n\n"
            f"👨‍⚕️ الأطباء: {doctors}\n"
            f"🏪 الصيدليات: {pharmacies}\n"
            f"🧑‍🔬 الصيادلة: {pharmacists}\n"
            f"💊 المنتجات: {products}\n"
            f"🧪 التشغيلات: {batches}\n"
            f"🏥 المستشفيات: {hospitals}\n"
            f"📦 المخازن: {warehouses}\n"
            f"👨‍💼 المستخدمون: {users}\n\n"
            f"🗓️ زيارات الأطباء اليوم: {doctor_visits}\n"
            f"🏪 زيارات الصيدليات اليوم: {pharmacy_visits}"
        )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# DOCTORS LIST
# ============================================================

async def doctors_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        doctors = db.query(
            Doctor
        ).order_by(
            Doctor.full_name
        ).limit(100).all()

        if not doctors:

            await query.message.reply_text(
                "لا يوجد أطباء مسجلون.\n\n"
                "استخدم زر «إضافة طبيب» لإضافة طبيب جديد."
            )

            return

        text = "👨‍⚕️ قائمة الأطباء:\n\n"

        for doctor in doctors:

            specialty = "-"

            if doctor.specialty:
                specialty = doctor.specialty.name

            text += (
                f"🆔 {doctor.id}\n"
                f"👨‍⚕️ {doctor.full_name}\n"
                f"🩺 {specialty}\n"
                f"⭐ {doctor.category}\n"
                f"📞 {doctor.phone or '-'}\n"
                f"📍 {doctor.city or '-'}\n\n"
            )

            if len(text) >= 3500:

                await query.message.reply_text(text)

                text = ""

        if text:
            await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# PHARMACIES LIST
# ============================================================

async def pharmacies_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        pharmacies = db.query(
            Pharmacy
        ).order_by(
            Pharmacy.name
        ).limit(100).all()

        if not pharmacies:

            await query.message.reply_text(
                "لا توجد صيدليات مسجلة.\n\n"
                "استخدم زر «إضافة صيدلية» لإضافة صيدلية."
            )

            return

        text = "🏪 قائمة الصيدليات:\n\n"

        for pharmacy in pharmacies:

            text += (
                f"🆔 {pharmacy.id}\n"
                f"🏪 {pharmacy.name}\n"
                f"👤 {pharmacy.owner_name or '-'}\n"
                f"📞 {pharmacy.phone or '-'}\n"
                f"📍 {pharmacy.city or '-'}\n"
                f"⭐ {pharmacy.classification or '-'}\n\n"
            )

            if len(text) >= 3500:

                await query.message.reply_text(text)

                text = ""

        if text:
            await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# PHARMACISTS
# ============================================================

async def pharmacists_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        pharmacists = db.query(
            Pharmacist
        ).order_by(
            Pharmacist.full_name
        ).limit(100).all()

        if not pharmacists:

            await query.message.reply_text(
                "لا يوجد صيادلة مسجلون حالياً."
            )

            return

        text = "🧑‍🔬 قائمة الصيادلة:\n\n"

        for pharmacist in pharmacists:

            pharmacy_name = "-"

            if pharmacist.pharmacy:
                pharmacy_name = pharmacist.pharmacy.name

            text += (
                f"🆔 {pharmacist.id}\n"
                f"🧑‍🔬 {pharmacist.full_name}\n"
                f"📞 {pharmacist.phone or '-'}\n"
                f"🏪 {pharmacy_name}\n"
                f"⭐ {pharmacist.classification or '-'}\n\n"
            )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# HOSPITALS
# ============================================================

async def hospitals_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        hospitals = db.query(
            Hospital
        ).order_by(
            Hospital.name
        ).limit(100).all()

        if not hospitals:

            await query.message.reply_text(
                "لا توجد مستشفيات مسجلة حالياً."
            )

            return

        text = "🏥 المستشفيات:\n\n"

        for hospital in hospitals:

            text += (
                f"🆔 {hospital.id}\n"
                f"🏥 {hospital.name}\n"
                f"📍 {hospital.city or '-'}\n"
                f"📞 {hospital.phone or '-'}\n"
                f"🏷️ {hospital.hospital_type or '-'}\n\n"
            )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# WAREHOUSES
# ============================================================

async def warehouses_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        warehouses = db.query(
            Warehouse
        ).order_by(
            Warehouse.name
        ).all()

        if not warehouses:

            await query.message.reply_text(
                "لا توجد مخازن مسجلة حالياً."
            )

            return

        text = "📦 المخازن:\n\n"

        for warehouse in warehouses:

            stock_count = db.query(
                func.coalesce(
                    func.sum(
                        WarehouseStock.quantity
                    ),
                    0
                )
            ).filter(
                WarehouseStock.warehouse_id ==
                warehouse.id
            ).scalar() or 0

            text += (
                f"🆔 {warehouse.id}\n"
                f"📦 {warehouse.name}\n"
                f"📍 {warehouse.city or '-'}\n"
                f"👤 {warehouse.manager_name or '-'}\n"
                f"📞 {warehouse.phone or '-'}\n"
                f"🔢 إجمالي الكميات: {stock_count}\n\n"
            )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# USERS
# ============================================================

async def users_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    if not is_admin(user_data):

        await query.message.reply_text(
            "❌ هذه الصفحة للمدير فقط."
        )

        return

    db = SessionLocal()

    try:

        users = db.query(
            User
        ).order_by(
            User.full_name
        ).all()

        if not users:

            await query.message.reply_text(
                "لا يوجد مستخدمون."
            )

            return

        text = "👨‍💼 المستخدمون:\n\n"

        for user in users:

            text += (
                f"🆔 {user.id}\n"
                f"👤 {user.full_name or '-'}\n"
                f"📱 @{user.username or '-'}\n"
                f"🔑 {user.role}\n"
                f"📞 {user.phone or '-'}\n"
                f"🟢 {'فعال' if user.is_active else 'غير فعال'}\n\n"
            )

            if len(text) >= 3500:

                await query.message.reply_text(text)

                text = ""

        if text:
            await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# PRODUCTS
# ============================================================

async def products_list_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        products = db.query(
            Product
        ).filter(
            Product.is_active == True
        ).order_by(
            Product.brand_name
        ).limit(200).all()

        if not products:

            await query.message.reply_text(
                "لا توجد منتجات حالياً.\n\n"
                "استخدم زر «إضافة منتج» لإضافة منتج."
            )

            return

        text = "💊 قائمة المنتجات:\n\n"

        for product in products:

            text += (
                f"🆔 {product.id}\n"
                f"🔢 الكود: {product.code}\n"
                f"💊 {product.brand_name}\n"
                f"🔬 {product.scientific_name or '-'}\n"
                f"⚗️ {product.concentration or '-'}\n"
                f"💉 {product.dosage_form or '-'}\n"
                f"🏭 {product.manufacturer or '-'}\n\n"
            )

            if len(text) >= 3500:

                await query.message.reply_text(text)

                text = ""

        if text:
            await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# ADD DOCTOR
# ============================================================

async def add_doctor_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    if not user_data:
        return ConversationHandler.END

    await query.message.reply_text(
        "👨‍⚕️ أدخل الاسم الكامل للطبيب:"
    )

    return ADD_DOCTOR_NAME


async def add_doctor_name(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data["doctor_name"] = (
        update.message.text.strip()
    )

    await update.message.reply_text(
        "📞 أدخل رقم هاتف الطبيب:"
    )

    return ADD_DOCTOR_PHONE


async def add_doctor_phone(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data["doctor_phone"] = (
        update.message.text.strip()
    )

    await update.message.reply_text(
        "🩺 أدخل تخصص الطبيب:\n\n"
        "مثال: أطفال"
    )

    return ADD_DOCTOR_SPECIALTY


async def add_doctor_specialty(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    name = update.message.text.strip()

    db = SessionLocal()

    try:

        specialty = db.query(
            Specialty
        ).filter(
            Specialty.name == name
        ).first()

        if not specialty:

            specialty = Specialty(name=name)

            db.add(specialty)

            db.commit()

            db.refresh(specialty)

        context.user_data[
            "specialty_id"
        ] = specialty.id

    finally:

        db.close()

    await update.message.reply_text(
        "⭐ اختر تصنيف الطبيب:\n"
        "A أو B أو C أو D"
    )

    return ADD_DOCTOR_CATEGORY


async def add_doctor_category(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    category = (
        update.message.text.strip().upper()
    )

    if category not in [
        "A",
        "B",
        "C",
        "D"
    ]:

        await update.message.reply_text(
            "❌ أدخل A أو B أو C أو D."
        )

        return ADD_DOCTOR_CATEGORY

    context.user_data[
        "doctor_category"
    ] = category

    await update.message.reply_text(
        "📍 أدخل المدينة أو المنطقة:"
    )

    return ADD_DOCTOR_CITY


async def add_doctor_city(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    city = update.message.text.strip()

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        doctor = Doctor(
            full_name=context.user_data.get(
                "doctor_name"
            ),
            phone=context.user_data.get(
                "doctor_phone"
            ),
            specialty_id=context.user_data.get(
                "specialty_id"
            ),
            category=context.user_data.get(
                "doctor_category"
            ),
            city=city,
            assigned_rep_id=user_data["id"]
        )

        db.add(doctor)

        db.commit()

        db.refresh(doctor)

        add_audit_log(
            db,
            user_data["id"],
            "إضافة طبيب",
            "doctor",
            doctor.id
        )

        db.commit()

        await update.message.reply_text(
            "✅ تم إضافة الطبيب بنجاح.\n\n"
            f"👨‍⚕️ {doctor.full_name}\n"
            f"🆔 {doctor.id}\n"
            f"⭐ {doctor.category}"
        )

    except Exception as error:

        db.rollback()

        logger.exception(error)

        await update.message.reply_text(
            "❌ حدث خطأ أثناء إضافة الطبيب."
        )

    finally:

        db.close()

    context.user_data.clear()

    return ConversationHandler.END


# ============================================================
# ADD PHARMACY
# ============================================================

async def add_pharmacy_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        "🏪 أدخل اسم الصيدلية:"
    )

    return ADD_PHARMACY_NAME


async def add_pharmacy_name(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "pharmacy_name"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "👤 أدخل اسم صاحب الصيدلية:"
    )

    return ADD_PHARMACY_OWNER


async def add_pharmacy_owner(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "pharmacy_owner"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "📞 أدخل رقم الهاتف:"
    )

    return ADD_PHARMACY_PHONE


async def add_pharmacy_phone(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "pharmacy_phone"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "📍 أدخل المدينة:"
    )

    return ADD_PHARMACY_CITY


async def add_pharmacy_city(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        pharmacy = Pharmacy(
            name=context.user_data.get(
                "pharmacy_name"
            ),
            owner_name=context.user_data.get(
                "pharmacy_owner"
            ),
            phone=context.user_data.get(
                "pharmacy_phone"
            ),
            city=update.message.text.strip(),
            assigned_rep_id=user_data["id"]
        )

        db.add(pharmacy)

        db.commit()

        db.refresh(pharmacy)

        add_audit_log(
            db,
            user_data["id"],
            "إضافة صيدلية",
            "pharmacy",
            pharmacy.id
        )

        db.commit()

        await update.message.reply_text(
            "✅ تم إضافة الصيدلية بنجاح.\n\n"
            f"🏪 {pharmacy.name}\n"
            f"🆔 {pharmacy.id}"
        )

    except Exception as error:

        db.rollback()

        logger.exception(error)

        await update.message.reply_text(
            "❌ حدث خطأ أثناء إضافة الصيدلية."
        )

    finally:

        db.close()

    context.user_data.clear()

    return ConversationHandler.END


# ============================================================
# ADD PRODUCT
# ============================================================

async def add_product_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        "🔢 أدخل كود المنتج:"
    )

    return ADD_PRODUCT_CODE


async def add_product_code(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "product_code"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "💊 أدخل الاسم التجاري:"
    )

    return ADD_PRODUCT_NAME


async def add_product_name(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "product_name"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "🔬 أدخل الاسم العلمي أو المادة الفعالة:"
    )

    return ADD_PRODUCT_SCIENTIFIC


async def add_product_scientific(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "product_scientific"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "⚗️ أدخل التركيز:"
    )

    return ADD_PRODUCT_CONCENTRATION


async def add_product_concentration(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "product_concentration"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "💉 أدخل الشكل الدوائي:"
    )

    return ADD_PRODUCT_FORM


async def add_product_form(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        code = context.user_data.get(
            "product_code"
        )

        existing = db.query(
            Product
        ).filter(
            Product.code == code
        ).first()

        if existing:

            await update.message.reply_text(
                "❌ هذا الكود مستخدم مسبقاً."
            )

            return ConversationHandler.END

        product = Product(
            code=code,
            brand_name=context.user_data.get(
                "product_name"
            ),
            scientific_name=context.user_data.get(
                "product_scientific"
            ),
            active_ingredient=context.user_data.get(
                "product_scientific"
            ),
            concentration=context.user_data.get(
                "product_concentration"
            ),
            dosage_form=update.message.text.strip()
        )

        db.add(product)

        db.commit()

        db.refresh(product)

        add_audit_log(
            db,
            user_data["id"],
            "إضافة منتج",
            "product",
            product.id
        )

        db.commit()

        await update.message.reply_text(
            "✅ تم إضافة المنتج.\n\n"
            f"💊 {product.brand_name}\n"
            f"🔢 {product.code}\n"
            f"🆔 {product.id}"
        )

    except Exception as error:

        db.rollback()

        logger.exception(error)

        await update.message.reply_text(
            "❌ حدث خطأ أثناء إضافة المنتج."
        )

    finally:

        db.close()

    context.user_data.clear()

    return ConversationHandler.END


# ============================================================
# USER PHARMACIES
# ============================================================

def get_user_pharmacies(user_id):

    db = SessionLocal()

    try:

        links = db.query(
            UserPharmacy
        ).filter(
            UserPharmacy.user_id == user_id
        ).all()

        result = []

        for link in links:

            pharmacy = db.query(
                Pharmacy
            ).filter(
                Pharmacy.id == link.pharmacy_id
            ).first()

            if pharmacy:
                result.append(pharmacy)

        return result

    finally:

        db.close()


# ============================================================
# MY PHARMACY
# ============================================================

async def my_pharmacy_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "⚠️ لم يتم ربط حسابك بأي صيدلية بعد."
        )

        return

    text = "🏪 صيدلياتي:\n\n"

    for pharmacy in pharmacies:

        text += (
            f"🆔 {pharmacy.id}\n"
            f"🏪 {pharmacy.name}\n"
            f"👤 {pharmacy.owner_name or '-'}\n"
            f"📞 {pharmacy.phone or '-'}\n"
            f"📍 {pharmacy.city or '-'}\n"
            f"⭐ {pharmacy.classification or '-'}\n\n"
        )

    await query.message.reply_text(text)


# ============================================================
# ADD PHARMACY STOCK
# ============================================================

async def pharmacy_add_stock_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة بحسابك."
        )

        return ConversationHandler.END

    context.user_data[
        "stock_pharmacy_id"
    ] = pharmacies[0].id

    await query.message.reply_text(
        "💊 أدخل اسم المنتج أو الكود:"
    )

    return ADD_STOCK_PRODUCT


async def pharmacy_add_stock_product(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    value = update.message.text.strip()

    db = SessionLocal()

    try:

        product = db.query(
            Product
        ).filter(
            or_(
                Product.code == value,
                Product.brand_name.ilike(
                    f"%{value}%"
                ),
                Product.scientific_name.ilike(
                    f"%{value}%"
                )
            )
        ).first()

        if not product:

            await update.message.reply_text(
                "❌ المنتج غير موجود."
            )

            return ADD_STOCK_PRODUCT

        context.user_data[
            "stock_product_id"
        ] = product.id

        await update.message.reply_text(
            f"💊 المنتج: {product.brand_name}\n\n"
            "🧪 أدخل رقم التشغيلة:"
        )

    finally:

        db.close()

    return ADD_STOCK_BATCH


async def pharmacy_add_stock_batch(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data[
        "stock_batch_number"
    ] = update.message.text.strip()

    await update.message.reply_text(
        "🔢 أدخل الكمية:"
    )

    return ADD_STOCK_QUANTITY


async def pharmacy_add_stock_quantity(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    try:

        quantity = int(
            update.message.text.strip()
        )

        if quantity < 0:
            raise ValueError

    except ValueError:

        await update.message.reply_text(
            "❌ أدخل كمية صحيحة."
        )

        return ADD_STOCK_QUANTITY

    context.user_data[
        "stock_quantity"
    ] = quantity

    await update.message.reply_text(
        "📅 أدخل تاريخ الانتهاء:\n"
        "YYYY-MM-DD"
    )

    return ADD_STOCK_EXPIRY


async def pharmacy_add_stock_expiry(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    try:

        expiry = datetime.strptime(
            update.message.text.strip(),
            "%Y-%m-%d"
        ).date()

    except ValueError:

        await update.message.reply_text(
            "❌ التاريخ غير صحيح."
        )

        return ADD_STOCK_EXPIRY

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        product_id = context.user_data[
            "stock_product_id"
        ]

        pharmacy_id = context.user_data[
            "stock_pharmacy_id"
        ]

        batch_number = context.user_data[
            "stock_batch_number"
        ]

        quantity = context.user_data[
            "stock_quantity"
        ]

        batch = db.query(
            Batch
        ).filter(
            Batch.product_id == product_id,
            Batch.batch_number == batch_number
        ).first()

        if not batch:

            batch = Batch(
                product_id=product_id,
                batch_number=batch_number,
                expiry_date=expiry
            )

            db.add(batch)

            db.commit()

            db.refresh(batch)

        elif not batch.expiry_date:

            batch.expiry_date = expiry

        stock = db.query(
            PharmacyStock
        ).filter(
            PharmacyStock.pharmacy_id == pharmacy_id,
            PharmacyStock.product_id == product_id,
            PharmacyStock.batch_id == batch.id
        ).first()

        if stock:

            before = stock.quantity

            stock.quantity += quantity

            stock.last_reported_by = user_data["id"]

            after = stock.quantity

        else:

            before = 0

            after = quantity

            stock = PharmacyStock(
                pharmacy_id=pharmacy_id,
                product_id=product_id,
                batch_id=batch.id,
                quantity=quantity,
                last_reported_by=user_data["id"]
            )

            db.add(stock)

            db.flush()

        db.add(
            StockMovement(
                location_type="pharmacy",
                location_id=pharmacy_id,
                product_id=product_id,
                batch_id=batch.id,
                movement_type="إضافة مخزون",
                quantity_before=before,
                quantity_change=quantity,
                quantity_after=after,
                user_id=user_data["id"]
            )
        )

        add_audit_log(
            db,
            user_data["id"],
            "تحديث مخزون الصيدلية",
            "pharmacy_stock",
            stock.id,
            str(before),
            str(after)
        )

        db.commit()

        product = db.query(
            Product
        ).filter(
            Product.id == product_id
        ).first()

        await update.message.reply_text(
            "✅ تم تحديث المخزون.\n\n"
            f"💊 {product.brand_name}\n"
            f"🧪 التشغيلة: {batch.batch_number}\n"
            f"🔢 الكمية: {after}\n"
            f"📅 الانتهاء: {expiry}"
        )

    except Exception as error:

        db.rollback()

        logger.exception(error)

        await update.message.reply_text(
            "❌ حدث خطأ أثناء حفظ المخزون."
        )

    finally:

        db.close()

    context.user_data.clear()

    return ConversationHandler.END


# ============================================================
# PHARMACY INVENTORY
# ============================================================

async def pharmacy_inventory_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة بحسابك."
        )

        return

    db = SessionLocal()

    try:

        text = "📦 مخزون الصيدلية:\n\n"

        for pharmacy in pharmacies:

            text += (
                f"🏪 {pharmacy.name}\n"
                "────────────────\n"
            )

            stocks = db.query(
                PharmacyStock
            ).filter(
                PharmacyStock.pharmacy_id ==
                pharmacy.id
            ).all()

            if not stocks:

                text += "لا توجد أصناف.\n\n"

            for stock in stocks:

                product = db.query(
                    Product
                ).filter(
                    Product.id == stock.product_id
                ).first()

                batch = None

                if stock.batch_id:

                    batch = db.query(
                        Batch
                    ).filter(
                        Batch.id == stock.batch_id
                    ).first()

                text += (
                    f"💊 {product.brand_name if product else '-'}\n"
                    f"🔢 الكمية: {stock.quantity}\n"
                    f"🧪 التشغيلة: "
                    f"{batch.batch_number if batch else '-'}\n"
                    f"📅 الانتهاء: "
                    f"{batch.expiry_date if batch else '-'}\n\n"
                )

                if len(text) >= 3500:

                    await query.message.reply_text(text)

                    text = ""

        if text:
            await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# LOW STOCK
# ============================================================

async def pharmacy_low_stock_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة."
        )

        return

    db = SessionLocal()

    try:

        text = "⚠️ المخزون المنخفض:\n\n"

        found = False

        for pharmacy in pharmacies:

            products = db.query(Product).filter(
                Product.is_active == True
            ).all()

            for product in products:

                quantity = db.query(
                    func.coalesce(
                        func.sum(
                            PharmacyStock.quantity
                        ),
                        0
                    )
                ).filter(
                    PharmacyStock.pharmacy_id ==
                    pharmacy.id,
                    PharmacyStock.product_id ==
                    product.id
                ).scalar() or 0

                if (
                    product.minimum_stock > 0
                    and quantity <= product.minimum_stock
                ):

                    found = True

                    text += (
                        f"🏪 {pharmacy.name}\n"
                        f"💊 {product.brand_name}\n"
                        f"🔢 الموجود: {quantity}\n"
                        f"⚠️ الحد الأدنى: "
                        f"{product.minimum_stock}\n\n"
                    )

        if not found:

            text += "🟢 لا توجد أصناف منخفضة المخزون."

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# EXPIRY
# ============================================================

async def pharmacy_expiry_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة."
        )

        return

    db = SessionLocal()

    try:

        today = date.today()

        warning = today + timedelta(days=180)

        text = "⏳ التشغيلات القريبة من الانتهاء:\n\n"

        found = False

        for pharmacy in pharmacies:

            stocks = db.query(
                PharmacyStock
            ).filter(
                PharmacyStock.pharmacy_id ==
                pharmacy.id
            ).all()

            for stock in stocks:

                if not stock.batch_id:
                    continue

                batch = db.query(
                    Batch
                ).filter(
                    Batch.id == stock.batch_id
                ).first()

                if not batch or not batch.expiry_date:
                    continue

                if batch.expiry_date <= warning:

                    product = db.query(
                        Product
                    ).filter(
                        Product.id == stock.product_id
                    ).first()

                    found = True

                    if batch.expiry_date < today:
                        status = "🔴 منتهية"
                    elif batch.expiry_date <= today + timedelta(days=30):
                        status = "🔴 أقل من شهر"
                    elif batch.expiry_date <= today + timedelta(days=90):
                        status = "🟠 أقل من 3 أشهر"
                    else:
                        status = "🟡 أقل من 6 أشهر"

                    text += (
                        f"{status}\n"
                        f"🏪 {pharmacy.name}\n"
                        f"💊 {product.brand_name if product else '-'}\n"
                        f"🧪 {batch.batch_number}\n"
                        f"📅 {batch.expiry_date}\n"
                        f"🔢 الكمية: {stock.quantity}\n\n"
                    )

        if not found:

            text += "🟢 لا توجد تشغيلات قريبة من الانتهاء."

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# ADMIN ALERTS
# ============================================================

async def alerts_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    db = SessionLocal()

    try:

        today = date.today()

        warning = today + timedelta(days=180)

        batches = db.query(
            Batch
        ).filter(
            Batch.expiry_date != None,
            Batch.expiry_date <= warning
        ).order_by(
            Batch.expiry_date
        ).all()

        if not batches:

            await query.message.reply_text(
                "🟢 لا توجد تنبيهات صلاحية."
            )

            return

        text = "⚠️ تنبيهات التشغيلات:\n\n"

        for batch in batches:

            product = db.query(
                Product
            ).filter(
                Product.id == batch.product_id
            ).first()

            if batch.expiry_date < today:
                status = "🔴 منتهي"
            elif batch.expiry_date <= today + timedelta(days=30):
                status = "🔴 أقل من شهر"
            elif batch.expiry_date <= today + timedelta(days=90):
                status = "🟠 أقل من 3 أشهر"
            else:
                status = "🟡 أقل من 6 أشهر"

            text += (
                f"{status}\n"
                f"💊 {product.brand_name if product else '-'}\n"
                f"🧪 {batch.batch_number}\n"
                f"📅 {batch.expiry_date}\n\n"
            )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# REPORTS
# ============================================================

async def reports_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    if not is_admin(user_data):

        await query.message.reply_text(
            "❌ التقرير العام متاح للمدير."
        )

        return

    db = SessionLocal()

    try:

        doctors = db.query(
            func.count(Doctor.id)
        ).scalar() or 0

        category_a = db.query(
            func.count(Doctor.id)
        ).filter(
            Doctor.category == "A"
        ).scalar() or 0

        category_b = db.query(
            func.count(Doctor.id)
        ).filter(
            Doctor.category == "B"
        ).scalar() or 0

        category_c = db.query(
            func.count(Doctor.id)
        ).filter(
            Doctor.category == "C"
        ).scalar() or 0

        category_d = db.query(
            func.count(Doctor.id)
        ).filter(
            Doctor.category == "D"
        ).scalar() or 0

        pharmacies = db.query(
            func.count(Pharmacy.id)
        ).scalar() or 0

        pharmacists = db.query(
            func.count(Pharmacist.id)
        ).scalar() or 0

        products = db.query(
            func.count(Product.id)
        ).scalar() or 0

        hospitals = db.query(
            func.count(Hospital.id)
        ).scalar() or 0

        warehouses = db.query(
            func.count(Warehouse.id)
        ).scalar() or 0

        batches = db.query(
            func.count(Batch.id)
        ).scalar() or 0

        visits = db.query(
            func.count(DoctorVisit.id)
        ).scalar() or 0

        pharmacy_visits = db.query(
            func.count(PharmacyVisit.id)
        ).scalar() or 0

        pending_requests = db.query(
            func.count(VisitRequest.id)
        ).filter(
            VisitRequest.status == "pending"
        ).scalar() or 0

        text = (
            "📊 التقرير العام للشركة\n\n"
            f"👨‍⚕️ الأطباء: {doctors}\n"
            f"  ⭐ A: {category_a}\n"
            f"  ⭐ B: {category_b}\n"
            f"  ⭐ C: {category_c}\n"
            f"  ⭐ D: {category_d}\n\n"
            f"🏪 الصيدليات: {pharmacies}\n"
            f"🧑‍🔬 الصيادلة: {pharmacists}\n"
            f"💊 المنتجات: {products}\n"
            f"🧪 التشغيلات: {batches}\n"
            f"🏥 المستشفيات: {hospitals}\n"
            f"📦 المخازن: {warehouses}\n\n"
            f"🗓️ زيارات الأطباء: {visits}\n"
            f"🏪 زيارات الصيدليات: {pharmacy_visits}\n"
            f"📋 طلبات الزيارة المعلقة: {pending_requests}"
        )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# REQUEST VISIT
# ============================================================

async def request_visit_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة بحسابك."
        )

        return

    db = SessionLocal()

    try:

        request = VisitRequest(
            pharmacy_id=pharmacies[0].id,
            requested_by_user_id=user_data["id"],
            status="pending",
            notes="طلب زيارة مندوب"
        )

        db.add(request)

        db.commit()

        await query.message.reply_text(
            "✅ تم إرسال طلب زيارة المندوب بنجاح."
        )

        # إشعار المديرين
        for admin_id in ADMIN_IDS:

            try:

                await context.bot.send_message(
                    chat_id=admin_id,
                    text=(
                        "🔔 طلب زيارة جديد\n\n"
                        f"🏪 الصيدلية: "
                        f"{pharmacies[0].name}\n"
                        f"👤 المستخدم: "
                        f"{user_data['full_name']}\n"
                        f"🆔 الطلب: {request.id}"
                    )
                )

            except Exception as error:

                logger.warning(
                    "تعذر إرسال إشعار للمدير: %s",
                    error
                )

    finally:

        db.close()


# ============================================================
# SEARCH
# ============================================================

async def search_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        "🔎 أدخل ما تريد البحث عنه:\n\n"
        "• اسم طبيب\n"
        "• اسم صيدلية\n"
        "• اسم منتج\n"
        "• اسم علمي\n"
        "• رقم هاتف\n"
        "• كود منتج\n"
        "• رقم تشغيلة"
    )

    return SEARCH_TEXT


async def search_execute(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    search = update.message.text.strip()

    db = SessionLocal()

    try:

        text = (
            f"🔎 نتائج البحث عن:\n"
            f"{search}\n\n"
        )

        # ----------------------------------------------------
        # Doctors
        # ----------------------------------------------------

        doctors = db.query(
            Doctor
        ).filter(
            or_(
                Doctor.full_name.ilike(
                    f"%{search}%"
                ),
                Doctor.phone.ilike(
                    f"%{search}%"
                )
            )
        ).limit(20).all()

        if doctors:

            text += "👨‍⚕️ الأطباء:\n"

            for doctor in doctors:

                specialty = "-"

                if doctor.specialty:
                    specialty = doctor.specialty.name

                text += (
                    f"• {doctor.full_name}\n"
                    f"  🩺 {specialty}\n"
                    f"  📞 {doctor.phone or '-'}\n"
                    f"  ⭐ {doctor.category}\n"
                    f"  📍 {doctor.city or '-'}\n\n"
                )

        # ----------------------------------------------------
        # Pharmacies
        # ----------------------------------------------------

        pharmacies = db.query(
            Pharmacy
        ).filter(
            or_(
                Pharmacy.name.ilike(
                    f"%{search}%"
                ),
                Pharmacy.phone.ilike(
                    f"%{search}%"
                ),
                Pharmacy.owner_name.ilike(
                    f"%{search}%"
                )
            )
        ).limit(20).all()

        if pharmacies:

            text += "🏪 الصيدليات:\n"

            for pharmacy in pharmacies:

                text += (
                    f"• {pharmacy.name}\n"
                    f"  👤 {pharmacy.owner_name or '-'}\n"
                    f"  📞 {pharmacy.phone or '-'}\n"
                    f"  📍 {pharmacy.city or '-'}\n\n"
                )

        # ----------------------------------------------------
        # Products
        # ----------------------------------------------------

        products = db.query(
            Product
        ).filter(
            or_(
                Product.brand_name.ilike(
                    f"%{search}%"
                ),
                Product.scientific_name.ilike(
                    f"%{search}%"
                ),
                Product.active_ingredient.ilike(
                    f"%{search}%"
                ),
                Product.code.ilike(
                    f"%{search}%"
                )
            )
        ).limit(20).all()

        if products:

            text += "💊 المنتجات:\n"

            for product in products:

                text += (
                    f"• {product.brand_name}\n"
                    f"  🔢 الكود: {product.code}\n"
                    f"  🔬 العلمي: "
                    f"{product.scientific_name or '-'}\n"
                    f"  ⚗️ التركيز: "
                    f"{product.concentration or '-'}\n\n"
                )

        # ----------------------------------------------------
        # Batches
        # ----------------------------------------------------

        batches = db.query(
            Batch
        ).filter(
            Batch.batch_number.ilike(
                f"%{search}%"
            )
        ).limit(20).all()

        if batches:

            text += "🧪 التشغيلات:\n"

            for batch in batches:

                product = db.query(
                    Product
                ).filter(
                    Product.id == batch.product_id
                ).first()

                text += (
                    f"• التشغيلة: "
                    f"{batch.batch_number}\n"
                    f"  💊 المنتج: "
                    f"{product.brand_name if product else '-'}\n"
                    f"  📅 الانتهاء: "
                    f"{batch.expiry_date or '-'}\n\n"
                )

        # ----------------------------------------------------
        # Pharmacists
        # ----------------------------------------------------

        pharmacists = db.query(
            Pharmacist
        ).filter(
            or_(
                Pharmacist.full_name.ilike(
                    f"%{search}%"
                ),
                Pharmacist.phone.ilike(
                    f"%{search}%"
                )
            )
        ).limit(20).all()

        if pharmacists:

            text += "🧑‍🔬 الصيادلة:\n"

            for pharmacist in pharmacists:

                pharmacy_name = "-"

                if pharmacist.pharmacy:
                    pharmacy_name = pharmacist.pharmacy.name

                text += (
                    f"• {pharmacist.full_name}\n"
                    f"  📞 {pharmacist.phone or '-'}\n"
                    f"  🏪 {pharmacy_name}\n\n"
                )

        if text.endswith(
            f"{search}\n\n"
        ):

            text += "❌ لم يتم العثور على نتائج."

        # Telegram limit
        chunks = []

        while len(text) > 3500:

            split_at = text.rfind(
                "\n",
                0,
                3500
            )

            if split_at <= 0:
                split_at = 3500

            chunks.append(
                text[:split_at]
            )

            text = text[split_at:]

        chunks.append(text)

        for chunk in chunks:

            if chunk.strip():

                await update.message.reply_text(
                    chunk
                )

    finally:

        db.close()

    return ConversationHandler.END


# ============================================================
# EXCEL HELPERS
# ============================================================

def excel_value(value):

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    if isinstance(value, date):
        return value.strftime(
            "%Y-%m-%d"
        )

    if isinstance(value, bool):
        return "نعم" if value else "لا"

    return value


def add_sheet(
    workbook,
    title,
    headers,
    rows
):

    ws = workbook.create_sheet(title)

    ws.append(headers)

    for cell in ws[1]:

        cell.font = cell.font.copy(
            bold=True
        )

    for row in rows:

        ws.append(
            [
                excel_value(value)
                for value in row
            ]
        )

    # Auto width
    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            value = str(
                cell.value or ""
            )

            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 10),
            50
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    return ws


# ============================================================
# BUILD EXCEL
# ============================================================

def build_excel_file(
    user_data
):

    db = SessionLocal()

    try:

        workbook = Workbook()

        default = workbook.active

        workbook.remove(default)

        # ----------------------------------------------------
        # USERS
        # ----------------------------------------------------

        users = db.query(User).all()

        add_sheet(
            workbook,
            "المستخدمون",
            [
                "ID",
                "Telegram ID",
                "الاسم",
                "Username",
                "الهاتف",
                "الصلاحية",
                "فعال",
                "تاريخ الإنشاء"
            ],
            [
                (
                    u.id,
                    u.telegram_id,
                    u.full_name,
                    u.username,
                    u.phone,
                    u.role,
                    u.is_active,
                    u.created_at
                )
                for u in users
            ]
        )

        # ----------------------------------------------------
        # DOCTORS
        # ----------------------------------------------------

        doctors = db.query(
            Doctor
        ).all()

        doctor_rows = []

        for d in doctors:

            specialty = (
                d.specialty.name
                if d.specialty
                else ""
            )

            doctor_rows.append(
                (
                    d.id,
                    d.full_name,
                    d.phone,
                    specialty,
                    d.category,
                    d.scientific_degree,
                    d.governorate,
                    d.city,
                    d.district,
                    d.address,
                    d.working_days,
                    d.working_hours,
                    d.importance_score,
                    d.prescription_score,
                    d.last_visit_date,
                    d.next_visit_date,
                    d.assigned_rep_id,
                    d.notes,
                    d.is_active,
                    d.created_at
                )
            )

        add_sheet(
            workbook,
            "الأطباء",
            [
                "ID",
                "الاسم",
                "الهاتف",
                "التخصص",
                "التصنيف",
                "الدرجة العلمية",
                "المحافظة",
                "المدينة",
                "المديرية",
                "العنوان",
                "أيام العمل",
                "ساعات العمل",
                "درجة الأهمية",
                "درجة الوصفات",
                "آخر زيارة",
                "الزيارة القادمة",
                "المندوب",
                "ملاحظات",
                "فعال",
                "تاريخ التسجيل"
            ],
            doctor_rows
        )

        # ----------------------------------------------------
        # PHARMACIES
        # ----------------------------------------------------

        pharmacies = db.query(
            Pharmacy
        ).all()

        add_sheet(
            workbook,
            "الصيدليات",
            [
                "ID",
                "اسم الصيدلية",
                "صاحب الصيدلية",
                "الهاتف",
                "المحافظة",
                "المدينة",
                "المديرية",
                "العنوان",
                "Latitude",
                "Longitude",
                "التصنيف",
                "المندوب",
                "آخر زيارة",
                "الزيارة القادمة",
                "ملاحظات",
                "فعال",
                "تاريخ التسجيل"
            ],
            [
                (
                    p.id,
                    p.name,
                    p.owner_name,
                    p.phone,
                    p.governorate,
                    p.city,
                    p.district,
                    p.address,
                    p.latitude,
                    p.longitude,
                    p.classification,
                    p.assigned_rep_id,
                    p.last_visit_date,
                    p.next_visit_date,
                    p.notes,
                    p.is_active,
                    p.created_at
                )
                for p in pharmacies
            ]
        )

        # ----------------------------------------------------
        # PHARMACISTS
        # ----------------------------------------------------

        pharmacists = db.query(
            Pharmacist
        ).all()

        pharmacist_rows = []

        for p in pharmacists:

            pharmacy_name = ""

            if p.pharmacy:
                pharmacy_name = p.pharmacy.name

            pharmacist_rows.append(
                (
                    p.id,
                    p.full_name,
                    p.phone,
                    p.pharmacy_id,
                    pharmacy_name,
                    p.classification,
                    p.notes
                )
            )

        add_sheet(
            workbook,
            "الصيادلة",
            [
                "ID",
                "الاسم",
                "الهاتف",
                "Pharmacy ID",
                "الصيدلية",
                "التصنيف",
                "ملاحظات"
            ],
            pharmacist_rows
        )

        # ----------------------------------------------------
        # PRODUCTS
        # ----------------------------------------------------

        products = db.query(
            Product
        ).all()

        add_sheet(
            workbook,
            "المنتجات",
            [
                "ID",
                "الكود",
                "الاسم التجاري",
                "الاسم العلمي",
                "المادة الفعالة",
                "التركيز",
                "الشكل الدوائي",
                "حجم العبوة",
                "الفئة العلاجية",
                "الشركة المصنعة",
                "بلد المنشأ",
                "السعر",
                "الحد الأدنى للمخزون",
                "الملاحظات العلمية",
                "فعال",
                "تاريخ التسجيل"
            ],
            [
                (
                    p.id,
                    p.code,
                    p.brand_name,
                    p.scientific_name,
                    p.active_ingredient,
                    p.concentration,
                    p.dosage_form,
                    p.package_size,
                    p.therapeutic_class,
                    p.manufacturer,
                    p.country_of_origin,
                    p.price,
                    p.minimum_stock,
                    p.scientific_notes,
                    p.is_active,
                    p.created_at
                )
                for p in products
            ]
        )

        # ----------------------------------------------------
        # BATCHES
        # ----------------------------------------------------

        batches = db.query(
            Batch
        ).all()

        batch_rows = []

        for b in batches:

            product_name = ""

            if b.product:
                product_name = b.product.brand_name

            batch_rows.append(
                (
                    b.id,
                    b.product_id,
                    product_name,
                    b.batch_number,
                    b.manufacture_date,
                    b.expiry_date
                )
            )

        add_sheet(
            workbook,
            "التشغيلات",
            [
                "ID",
                "Product ID",
                "المنتج",
                "رقم التشغيلة",
                "تاريخ التصنيع",
                "تاريخ الانتهاء"
            ],
            batch_rows
        )

        # ----------------------------------------------------
        # HOSPITALS
        # ----------------------------------------------------

        hospitals = db.query(
            Hospital
        ).all()

        add_sheet(
            workbook,
            "المستشفيات",
            [
                "ID",
                "الاسم",
                "النوع",
                "المحافظة",
                "المدينة",
                "المديرية",
                "العنوان",
                "الهاتف",
                "ملاحظات"
            ],
            [
                (
                    h.id,
                    h.name,
                    h.hospital_type,
                    h.governorate,
                    h.city,
                    h.district,
                    h.address,
                    h.phone,
                    h.notes
                )
                for h in hospitals
            ]
        )

        # ----------------------------------------------------
        # WAREHOUSES
        # ----------------------------------------------------

        warehouses = db.query(
            Warehouse
        ).all()

        add_sheet(
            workbook,
            "المخازن",
            [
                "ID",
                "الاسم",
                "المحافظة",
                "المدينة",
                "العنوان",
                "المدير",
                "الهاتف",
                "فعال"
            ],
            [
                (
                    w.id,
                    w.name,
                    w.governorate,
                    w.city,
                    w.address,
                    w.manager_name,
                    w.phone,
                    w.is_active
                )
                for w in warehouses
            ]
        )

        # ----------------------------------------------------
        # WAREHOUSE STOCK
        # ----------------------------------------------------

        warehouse_stock = db.query(
            WarehouseStock
        ).all()

        warehouse_stock_rows = []

        for s in warehouse_stock:

            warehouse_name = (
                s.warehouse.name
                if s.warehouse
                else ""
            )

            product_name = (
                s.product.brand_name
                if s.product
                else ""
            )

            batch_number = (
                s.batch.batch_number
                if s.batch
                else ""
            )

            warehouse_stock_rows.append(
                (
                    s.id,
                    s.warehouse_id,
                    warehouse_name,
                    s.product_id,
                    product_name,
                    s.batch_id,
                    batch_number,
                    s.quantity,
                    s.reserved_quantity,
                    s.updated_at
                )
            )

        add_sheet(
            workbook,
            "مخزون المخازن",
            [
                "ID",
                "Warehouse ID",
                "المخزن",
                "Product ID",
                "المنتج",
                "Batch ID",
                "التشغيلة",
                "الكمية",
                "الكمية المحجوزة",
                "آخر تحديث"
            ],
            warehouse_stock_rows
        )

        # ----------------------------------------------------
        # PHARMACY STOCK
        # ----------------------------------------------------

        pharmacy_stock = db.query(
            PharmacyStock
        ).all()

        pharmacy_stock_rows = []

        for s in pharmacy_stock:

            pharmacy_name = (
                s.pharmacy.name
                if s.pharmacy
                else ""
            )

            product_name = (
                s.product.brand_name
                if s.product
                else ""
            )

            batch_number = (
                s.batch.batch_number
                if s.batch
                else ""
            )

            pharmacy_stock_rows.append(
                (
                    s.id,
                    s.pharmacy_id,
                    pharmacy_name,
                    s.product_id,
                    product_name,
                    s.batch_id,
                    batch_number,
                    s.quantity,
                    s.last_reported_by,
                    s.updated_at
                )
            )

        add_sheet(
            workbook,
            "مخزون الصيدليات",
            [
                "ID",
                "Pharmacy ID",
                "الصيدلية",
                "Product ID",
                "المنتج",
                "Batch ID",
                "التشغيلة",
                "الكمية",
                "تم التحديث بواسطة",
                "آخر تحديث"
            ],
            pharmacy_stock_rows
        )

        # ----------------------------------------------------
        # STOCK MOVEMENTS
        # ----------------------------------------------------

        movements = db.query(
            StockMovement
        ).order_by(
            StockMovement.created_at.desc()
        ).all()

        add_sheet(
            workbook,
            "حركات المخزون",
            [
                "ID",
                "نوع الموقع",
                "Location ID",
                "Product ID",
                "Batch ID",
                "نوع الحركة",
                "قبل",
                "التغيير",
                "بعد",
                "User ID",
                "ملاحظات",
                "التاريخ"
            ],
            [
                (
                    m.id,
                    m.location_type,
                    m.location_id,
                    m.product_id,
                    m.batch_id,
                    m.movement_type,
                    m.quantity_before,
                    m.quantity_change,
                    m.quantity_after,
                    m.user_id,
                    m.notes,
                    m.created_at
                )
                for m in movements
            ]
        )

        # ----------------------------------------------------
        # DOCTOR VISITS
        # ----------------------------------------------------

        visits = db.query(
            DoctorVisit
        ).order_by(
            DoctorVisit.visit_date.desc()
        ).all()

        visit_rows = []

        for v in visits:

            doctor_name = (
                v.doctor.full_name
                if v.doctor
                else ""
            )

            visit_rows.append(
                (
                    v.id,
                    v.doctor_id,
                    doctor_name,
                    v.representative_id,
                    v.visit_date,
                    v.interest_level,
                    v.notes,
                    v.next_visit_date
                )
            )

        add_sheet(
            workbook,
            "زيارات الأطباء",
            [
                "ID",
                "Doctor ID",
                "الطبيب",
                "المندوب",
                "تاريخ الزيارة",
                "مستوى الاهتمام",
                "ملاحظات",
                "الزيارة القادمة"
            ],
            visit_rows
        )

        # ----------------------------------------------------
        # PHARMACY VISITS
        # ----------------------------------------------------

        pharmacy_visits = db.query(
            PharmacyVisit
        ).order_by(
            PharmacyVisit.visit_date.desc()
        ).all()

        pharmacy_visit_rows = []

        for v in pharmacy_visits:

            pharmacy_name = ""

            if v.pharmacy:
                pharmacy_name = v.pharmacy.name

            pharmacy_visit_rows.append(
                (
                    v.id,
                    v.pharmacy_id,
                    pharmacy_name,
                    v.representative_id,
                    v.visit_date,
                    v.notes,
                    v.next_visit_date
                )
            )

        add_sheet(
            workbook,
            "زيارات الصيدليات",
            [
                "ID",
                "Pharmacy ID",
                "الصيدلية",
                "المندوب",
                "تاريخ الزيارة",
                "ملاحظات",
                "الزيارة القادمة"
            ],
            pharmacy_visit_rows
        )

        # ----------------------------------------------------
        # VISIT REQUESTS
        # ----------------------------------------------------

        requests = db.query(
            VisitRequest
        ).order_by(
            VisitRequest.created_at.desc()
        ).all()

        request_rows = []

        for r in requests:

            pharmacy_name = ""

            if r.pharmacy:
                pharmacy_name = r.pharmacy.name

            request_rows.append(
                (
                    r.id,
                    r.pharmacy_id,
                    pharmacy_name,
                    r.requested_by_user_id,
                    r.status,
                    r.notes,
                    r.created_at
                )
            )

        add_sheet(
            workbook,
            "طلبات الزيارات",
            [
                "ID",
                "Pharmacy ID",
                "الصيدلية",
                "طلب بواسطة",
                "الحالة",
                "ملاحظات",
                "تاريخ الطلب"
            ],
            request_rows
        )

        # ----------------------------------------------------
        # REMINDERS
        # ----------------------------------------------------

        reminders = db.query(
            Reminder
        ).all()

        add_sheet(
            workbook,
            "التذكيرات",
            [
                "ID",
                "User ID",
                "العنوان",
                "الرسالة",
                "تاريخ التذكير",
                "تم الإرسال"
            ],
            [
                (
                    r.id,
                    r.user_id,
                    r.title,
                    r.message,
                    r.reminder_date,
                    r.is_sent
                )
                for r in reminders
            ]
        )

        # ----------------------------------------------------
        # AUDIT LOG
        # ----------------------------------------------------

        logs = db.query(
            AuditLog
        ).order_by(
            AuditLog.created_at.desc()
        ).all()

        add_sheet(
            workbook,
            "سجل العمليات",
            [
                "ID",
                "User ID",
                "العملية",
                "نوع البيانات",
                "Entity ID",
                "القيمة السابقة",
                "القيمة الجديدة",
                "التاريخ"
            ],
            [
                (
                    l.id,
                    l.user_id,
                    l.action,
                    l.entity_type,
                    l.entity_id,
                    l.old_value,
                    l.new_value,
                    l.created_at
                )
                for l in logs
            ]
        )

        # ----------------------------------------------------
        # SUMMARY
        # ----------------------------------------------------

        summary = workbook.create_sheet(
            "ملخص النظام",
            0
        )

        summary.append(
            [
                "المؤشر",
                "العدد"
            ]
        )

        summary_data = [
            (
                "الأطباء",
                db.query(func.count(Doctor.id)).scalar() or 0
            ),
            (
                "الصيدليات",
                db.query(func.count(Pharmacy.id)).scalar() or 0
            ),
            (
                "الصيادلة",
                db.query(func.count(Pharmacist.id)).scalar() or 0
            ),
            (
                "المنتجات",
                db.query(func.count(Product.id)).scalar() or 0
            ),
            (
                "التشغيلات",
                db.query(func.count(Batch.id)).scalar() or 0
            ),
            (
                "المستشفيات",
                db.query(func.count(Hospital.id)).scalar() or 0
            ),
            (
                "المخازن",
                db.query(func.count(Warehouse.id)).scalar() or 0
            ),
            (
                "المستخدمون",
                db.query(func.count(User.id)).scalar() or 0
            ),
            (
                "زيارات الأطباء",
                db.query(func.count(DoctorVisit.id)).scalar() or 0
            ),
            (
                "زيارات الصيدليات",
                db.query(func.count(PharmacyVisit.id)).scalar() or 0
            ),
            (
                "طلبات الزيارة",
                db.query(func.count(VisitRequest.id)).scalar() or 0
            ),
        ]

        for row in summary_data:
            summary.append(row)

        for cell in summary[1]:
            cell.font = cell.font.copy(
                bold=True
            )

        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 20

        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output

    finally:

        db.close()


# ============================================================
# EXPORT EXCEL
# ============================================================

async def export_excel_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer(
        "جاري تجهيز ملف Excel..."
    )

    user_data = get_or_create_user(update)

    if not user_data:

        return

    try:

        excel_file = build_excel_file(
            user_data
        )

        filename = (
            "medical_company_data_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        await query.message.reply_document(
            document=excel_file,
            filename=filename,
            caption=(
                "📊 تم استخراج بيانات النظام بالكامل إلى Excel.\n\n"
                "يحتوي الملف على:\n"
                "• ملخص النظام\n"
                "• الأطباء\n"
                "• الصيدليات\n"
                "• الصيادلة\n"
                "• المنتجات\n"
                "• التشغيلات\n"
                "• المستشفيات\n"
                "• المخازن\n"
                "• مخزون المخازن\n"
                "• مخزون الصيدليات\n"
                "• حركات المخزون\n"
                "• زيارات الأطباء\n"
                "• زيارات الصيدليات\n"
                "• طلبات الزيارات\n"
                "• المستخدمين\n"
                "• التذكيرات\n"
                "• سجل العمليات"
            )
        )

    except Exception as error:

        logger.exception(error)

        await query.message.reply_text(
            "❌ حدث خطأ أثناء إنشاء ملف Excel."
        )


# ============================================================
# TODAY VISITS
# ============================================================

async def today_visits_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        today = date.today()

        doctor_visits = db.query(
            DoctorVisit
        ).filter(
            DoctorVisit.visit_date == today
        ).all()

        pharmacy_visits = db.query(
            PharmacyVisit
        ).filter(
            PharmacyVisit.visit_date == today
        ).all()

        text = (
            f"🗓️ زيارات اليوم\n"
            f"📅 {today}\n\n"
        )

        text += "👨‍⚕️ زيارات الأطباء:\n"

        if doctor_visits:

            for visit in doctor_visits:

                doctor = visit.doctor

                text += (
                    f"• {doctor.full_name if doctor else '-'}\n"
                    f"  📝 {visit.notes or '-'}\n\n"
                )

        else:

            text += "لا توجد زيارات أطباء.\n\n"

        text += "🏪 زيارات الصيدليات:\n"

        if pharmacy_visits:

            for visit in pharmacy_visits:

                text += (
                    f"• الصيدلية رقم {visit.pharmacy_id}\n"
                    f"  📝 {visit.notes or '-'}\n\n"
                )

        else:

            text += "لا توجد زيارات صيدليات."

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# MY REMINDERS
# ============================================================

async def my_reminders_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    db = SessionLocal()

    try:

        reminders = db.query(
            Reminder
        ).filter(
            Reminder.user_id == user_data["id"],
            Reminder.is_sent == False
        ).order_by(
            Reminder.reminder_date
        ).all()

        if not reminders:

            await query.message.reply_text(
                "🔔 لا توجد تذكيرات حالياً."
            )

            return

        text = "🔔 تذكيراتي:\n\n"

        for reminder in reminders:

            text += (
                f"📌 {reminder.title}\n"
                f"📝 {reminder.message or '-'}\n"
                f"📅 {reminder.reminder_date}\n\n"
            )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# PHARMACY PRODUCTS
# ============================================================

async def pharmacy_products_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    user_data = get_or_create_user(update)

    pharmacies = get_user_pharmacies(
        user_data["id"]
    )

    if not pharmacies:

        await query.message.reply_text(
            "❌ لا توجد صيدلية مرتبطة."
        )

        return

    db = SessionLocal()

    try:

        text = "📋 قائمة أصناف الصيدلية:\n\n"

        for pharmacy in pharmacies:

            stocks = db.query(
                PharmacyStock
            ).filter(
                PharmacyStock.pharmacy_id ==
                pharmacy.id
            ).all()

            text += (
                f"🏪 {pharmacy.name}\n\n"
            )

            for stock in stocks:

                product = stock.product

                if product:

                    text += (
                        f"💊 {product.brand_name}\n"
                        f"🔢 {stock.quantity}\n"
                        f"🧪 "
                        f"{stock.batch.batch_number if stock.batch else '-'}\n\n"
                    )

        await query.message.reply_text(text)

    finally:

        db.close()


# ============================================================
# PHARMACY NOTE
# ============================================================

async def pharmacy_note_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        "💬 لإرسال ملاحظة، أرسلها في رسالة منفصلة إلى الإدارة."
    )


# ============================================================
# PLACEHOLDER STOCK UPDATE
# ============================================================

async def pharmacy_update_stock_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        "✏️ تحديث كمية المخزون:\n\n"
        "يمكنك استخدام «➕ إضافة صنف» لإضافة الكمية الجديدة "
        "وسيتم دمجها تلقائياً مع كمية الصنف الموجودة."
    )


# ============================================================
# CANCEL
# ============================================================

async def cancel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data.clear()

    await update.message.reply_text(
        "❌ تم إلغاء العملية."
    )

    return ConversationHandler.END


# ============================================================
# LINK PHARMACY
# ============================================================

async def add_pharmacy_user_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_data = get_or_create_user(update)

    if not is_admin(user_data):

        await update.message.reply_text(
            "❌ ليس لديك صلاحية."
        )

        return

    if len(context.args) < 2:

        await update.message.reply_text(
            "الاستخدام:\n"
            "/link_pharmacy USER_ID PHARMACY_ID"
        )

        return

    try:

        target_user_id = int(
            context.args[0]
        )

        pharmacy_id = int(
            context.args[1]
        )

    except ValueError:

        await update.message.reply_text(
            "❌ يجب أن تكون الأرقام صحيحة."
        )

        return

    db = SessionLocal()

    try:

        user = db.query(User).filter(
            User.id == target_user_id
        ).first()

        pharmacy = db.query(
            Pharmacy
        ).filter(
            Pharmacy.id == pharmacy_id
        ).first()

        if not user:

            await update.message.reply_text(
                "❌ المستخدم غير موجود."
            )

            return

        if not pharmacy:

            await update.message.reply_text(
                "❌ الصيدلية غير موجودة."
            )

            return

        existing = db.query(
            UserPharmacy
        ).filter(
            UserPharmacy.user_id == user.id,
            UserPharmacy.pharmacy_id == pharmacy.id
        ).first()

        if existing:

            await update.message.reply_text(
                "ℹ️ الربط موجود مسبقاً."
            )

            return

        db.add(
            UserPharmacy(
                user_id=user.id,
                pharmacy_id=pharmacy.id
            )
        )

        db.commit()

        await update.message.reply_text(
            "✅ تم ربط المستخدم بالصيدلية بنجاح."
        )

    except Exception as error:

        db.rollback()

        logger.exception(error)

        await update.message.reply_text(
            "❌ حدث خطأ أثناء الربط."
        )

    finally:

        db.close()


# ============================================================
# SPECIALTIES SEED
# ============================================================

def seed_specialties():

    specialties = [
        "طب عام",
        "الباطنية",
        "الأطفال",
        "النساء والولادة",
        "القلب",
        "الجراحة",
        "العظام",
        "الأنف والأذن والحنجرة",
        "الجلدية",
        "العيون",
        "المسالك البولية",
        "الأعصاب",
        "الأورام",
        "الأسنان",
        "الغدد الصماء",
        "الكلى",
        "الجهاز الهضمي",
        "الصدرية",
        "الطوارئ",
        "التخدير",
        "الأشعة",
        "المختبرات"
    ]

    db = SessionLocal()

    try:

        for name in specialties:

            existing = db.query(
                Specialty
            ).filter(
                Specialty.name == name
            ).first()

            if not existing:

                db.add(
                    Specialty(name=name)
                )

        db.commit()

    finally:

        db.close()


# ============================================================
# GENERAL MENU CALLBACK
# ============================================================

async def menu_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    data = query.data

    # يجب ألا نترك أي زر بدون تنفيذ

    if data == "admin_doctors":
        await doctors_list_callback(update, context)

    elif data == "admin_pharmacies":
        await pharmacies_list_callback(update, context)

    elif data == "admin_products":
        await products_list_callback(update, context)

    elif data == "admin_pharmacists":
        await pharmacists_list_callback(update, context)

    elif data == "admin_hospitals":
        await hospitals_list_callback(update, context)

    elif data == "admin_warehouses":
        await warehouses_list_callback(update, context)

    elif data == "admin_users":
        await users_list_callback(update, context)

    elif data == "rep_doctors":
        await doctors_list_callback(update, context)

    elif data == "rep_pharmacies":
        await pharmacies_list_callback(update, context)

    elif data == "products_list":
        await products_list_callback(update, context)

    elif data == "dashboard":
        await dashboard_callback(update, context)

    elif data == "admin_reports":
        await reports_callback(update, context)

    elif data == "admin_alerts":
        await alerts_callback(update, context)

    elif data == "search":
        # ConversationHandler هو المسؤول
        await query.answer()

    elif data == "pharmacy_inventory":
        await pharmacy_inventory_callback(update, context)

    elif data == "my_pharmacy":
        await my_pharmacy_callback(update, context)

    elif data == "pharmacy_products":
        await pharmacy_products_callback(update, context)

    elif data == "pharmacy_low_stock":
        await pharmacy_low_stock_callback(update, context)

    elif data == "pharmacy_expiry":
        await pharmacy_expiry_callback(update, context)

    elif data == "pharmacy_update_stock":
        await pharmacy_update_stock_callback(update, context)

    elif data == "request_visit":
        await request_visit_callback(update, context)

    elif data == "my_reminders":
        await my_reminders_callback(update, context)

    elif data == "today_visits":
        await today_visits_callback(update, context)

    elif data == "export_excel":
        await export_excel_callback(update, context)

    elif data == "pharmacy_note":
        await pharmacy_note_callback(update, context)

    else:

        await query.answer(
            "الخيار غير معروف."
        )


# ============================================================
# ERROR HANDLER
# ============================================================

async def error_handler(
    update: object,
    context: ContextTypes.DEFAULT_TYPE
):

    logger.exception(
        "حدث خطأ:",
        exc_info=context.error
    )


# ============================================================
# MAIN
# ============================================================

def main():

    initialize_database()

    seed_specialties()

    application = (
        Application.builder()
        .token(BOT_TOKEN)
        .build()
    )

    # ========================================================
    # START
    # ========================================================

    application.add_handler(
        CommandHandler(
            "start",
            start
        )
    )

    # ========================================================
    # DOCTOR CONVERSATION
    # ========================================================

    doctor_conversation = ConversationHandler(

        entry_points=[
            CallbackQueryHandler(
                add_doctor_start,
                pattern="^add_doctor$"
            )
        ],

        states={

            ADD_DOCTOR_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_doctor_name
                )
            ],

            ADD_DOCTOR_PHONE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_doctor_phone
                )
            ],

            ADD_DOCTOR_SPECIALTY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_doctor_specialty
                )
            ],

            ADD_DOCTOR_CATEGORY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_doctor_category
                )
            ],

            ADD_DOCTOR_CITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_doctor_city
                )
            ]

        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]

    )

    application.add_handler(
        doctor_conversation
    )

    # ========================================================
    # PHARMACY CONVERSATION
    # ========================================================

    pharmacy_conversation = ConversationHandler(

        entry_points=[
            CallbackQueryHandler(
                add_pharmacy_start,
                pattern="^add_pharmacy$"
            )
        ],

        states={

            ADD_PHARMACY_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_pharmacy_name
                )
            ],

            ADD_PHARMACY_OWNER: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_pharmacy_owner
                )
            ],

            ADD_PHARMACY_PHONE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_pharmacy_phone
                )
            ],

            ADD_PHARMACY_CITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_pharmacy_city
                )
            ]

        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]

    )

    application.add_handler(
        pharmacy_conversation
    )

    # ========================================================
    # PRODUCT CONVERSATION
    # ========================================================

    product_conversation = ConversationHandler(

        entry_points=[
            CallbackQueryHandler(
                add_product_start,
                pattern="^add_product$"
            )
        ],

        states={

            ADD_PRODUCT_CODE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_code
                )
            ],

            ADD_PRODUCT_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_name
                )
            ],

            ADD_PRODUCT_SCIENTIFIC: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_scientific
                )
            ],

            ADD_PRODUCT_CONCENTRATION: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_concentration
                )
            ],

            ADD_PRODUCT_FORM: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_form
                )
            ]

        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]

    )

    application.add_handler(
        product_conversation
    )

    # ========================================================
    # STOCK CONVERSATION
    # ========================================================

    stock_conversation = ConversationHandler(

        entry_points=[
            CallbackQueryHandler(
                pharmacy_add_stock_start,
                pattern="^pharmacy_add_stock$"
            )
        ],

        states={

            ADD_STOCK_PRODUCT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    pharmacy_add_stock_product
                )
            ],

            ADD_STOCK_BATCH: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    pharmacy_add_stock_batch
                )
            ],

            ADD_STOCK_QUANTITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    pharmacy_add_stock_quantity
                )
            ],

            ADD_STOCK_EXPIRY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    pharmacy_add_stock_expiry
                )
            ]

        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]

    )

    application.add_handler(
        stock_conversation
    )

    # ========================================================
    # SEARCH CONVERSATION
    # ========================================================

    search_conversation = ConversationHandler(

        entry_points=[
            CallbackQueryHandler(
                search_start,
                pattern="^search$"
            )
        ],

        states={

            SEARCH_TEXT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    search_execute
                )
            ]

        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]

    )

    application.add_handler(
        search_conversation
    )

    # ========================================================
    # DIRECT CALLBACKS
    # ========================================================

    application.add_handler(
        CallbackQueryHandler(
            dashboard_callback,
            pattern="^dashboard$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            export_excel_callback,
            pattern="^export_excel$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            today_visits_callback,
            pattern="^today_visits$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            my_reminders_callback,
            pattern="^my_reminders$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacy_inventory_callback,
            pattern="^pharmacy_inventory$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacy_products_callback,
            pattern="^pharmacy_products$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacy_low_stock_callback,
            pattern="^pharmacy_low_stock$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacy_expiry_callback,
            pattern="^pharmacy_expiry$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacy_update_stock_callback,
            pattern="^pharmacy_update_stock$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            my_pharmacy_callback,
            pattern="^my_pharmacy$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            request_visit_callback,
            pattern="^request_visit$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            doctors_list_callback,
            pattern="^(admin_doctors|rep_doctors)$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacies_list_callback,
            pattern="^(admin_pharmacies|rep_pharmacies)$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            products_list_callback,
            pattern="^(admin_products|products_list)$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            pharmacists_list_callback,
            pattern="^admin_pharmacists$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            hospitals_list_callback,
            pattern="^admin_hospitals$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            warehouses_list_callback,
            pattern="^admin_warehouses$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            users_list_callback,
            pattern="^admin_users$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            reports_callback,
            pattern="^admin_reports$"
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            alerts_callback,
            pattern="^admin_alerts$"
        )
    )

    # ========================================================
    # GENERAL FALLBACK CALLBACK
    # ========================================================

    application.add_handler(
        CallbackQueryHandler(
            menu_callback
        )
    )

    # ========================================================
    # COMMANDS
    # ========================================================

    application.add_handler(
        CommandHandler(
            "cancel",
            cancel
        )
    )

    application.add_handler(
        CommandHandler(
            "link_pharmacy",
            add_pharmacy_user_command
        )
    )

    # ========================================================
    # ERRORS
    # ========================================================

    application.add_error_handler(
        error_handler
    )

    logger.info(
        "Medical Representative Bot Started"
    )

    application.run_polling(
        allowed_updates=Update.ALL_TYPES
    )


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    main()
